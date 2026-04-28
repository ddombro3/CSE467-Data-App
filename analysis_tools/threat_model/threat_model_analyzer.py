from pathlib import Path
import argparse
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from threat_rules import (
    get_category_rule,
    get_risk_score,
    add_contextual_threats,
    needs_manual_review,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]

DEFAULT_INPUT_FILE = PROJECT_ROOT / "data" / "output" / "master_data_privacy_output.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "threat_output"

THREAT_OUTPUT_COLUMNS = [
    "Platform",
    "Source_Name",
    "Source URL",
    "Source_Type",
    "Evidence_Quote",
    "Data_Type",
    "Category",
    "Collection_Type",
    "Consumer_or_Api",
    "Sensitivity",
    "LINDDUN_Threats",
    "PANOPTIC_Activity",
    "Risk_Level",
    "Likelihood",
    "Consequence",
    "Risk_Score",
    "Threat_Reason",
    "Recommended_Mitigation",
    "Needs_Manual_Review",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Analyze extracted LLM privacy evidence using LINDDUN and PANOPTIC-style threat categories."
    )

    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT_FILE),
        help="Path to Project 1 master_data_privacy_output.xlsx file."
    )

    return parser.parse_args()


def read_input_file(input_path):
    path = Path(input_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Input file not found: {path}\n"
            "Run Project 1 first with: python src/main.py"
        )

    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return pd.read_excel(path)

    raise ValueError("Input file must be .csv, .xlsx, or .xlsm")


def build_threat_row(row):
    category = row.get("Category", "Unknown")
    rule = get_category_rule(category)

    linddun, panoptic = add_contextual_threats(
        row=row,
        base_linddun=rule["linddun"],
        base_panoptic=rule["panoptic_activity"],
    )

    risk_level = rule["risk_level"]
    likelihood, consequence, risk_score = get_risk_score(risk_level)

    return {
        "Platform": row.get("Platform"),
        "Source_Name": row.get("Source_Name"),
        "Source URL": row.get("Source URL"),
        "Source_Type": row.get("Source_Type"),
        "Evidence_Quote": row.get("Evidence_Quote"),
        "Data_Type": row.get("Data_Type"),
        "Category": category,
        "Collection_Type": row.get("Collection_Type"),
        "Consumer_or_Api": row.get("Consumer_or_Api"),
        "Sensitivity": row.get("Sensitivity"),
        "LINDDUN_Threats": ", ".join(linddun),
        "PANOPTIC_Activity": panoptic,
        "Risk_Level": risk_level,
        "Likelihood": likelihood,
        "Consequence": consequence,
        "Risk_Score": risk_score,
        "Threat_Reason": rule["reason"],
        "Recommended_Mitigation": rule["mitigation"],
        "Needs_Manual_Review": needs_manual_review(row, risk_level),
    }


def analyze_threats(df):
    threat_rows = []

    for _, row in df.iterrows():
        threat_rows.append(build_threat_row(row))

    return pd.DataFrame(threat_rows, columns=THREAT_OUTPUT_COLUMNS)


def build_summary(threat_df):
    if threat_df.empty:
        return pd.DataFrame()

    summary = (
        threat_df
        .groupby(["Risk_Level", "Category", "LINDDUN_Threats", "PANOPTIC_Activity"])
        .size()
        .reset_index(name="Count")
        .sort_values(by=["Risk_Level", "Count"], ascending=[True, False])
    )

    return summary


def format_excel_file(excel_path, sheet_name):
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    worksheet.title = sheet_name

    widths = {
        "A": 18,
        "B": 34,
        "C": 55,
        "D": 22,
        "E": 90,
        "F": 24,
        "G": 30,
        "H": 28,
        "I": 30,
        "J": 20,
        "K": 42,
        "L": 52,
        "M": 18,
        "N": 14,
        "O": 14,
        "P": 14,
        "Q": 65,
        "R": 65,
        "S": 22,
    }

    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[1].height = 45

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(excel_path)


def save_outputs(threat_df, summary_df):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    threat_csv = OUTPUT_DIR / "privacy_threat_model.csv"
    threat_xlsx = OUTPUT_DIR / "privacy_threat_model.xlsx"

    summary_csv = OUTPUT_DIR / "privacy_threat_summary.csv"
    summary_xlsx = OUTPUT_DIR / "privacy_threat_summary.xlsx"

    threat_df.to_csv(threat_csv, index=False)
    threat_df.to_excel(threat_xlsx, index=False)
    format_excel_file(threat_xlsx, "Privacy_Threat_Model")

    summary_df.to_csv(summary_csv, index=False)
    summary_df.to_excel(summary_xlsx, index=False)

    return threat_csv, threat_xlsx, summary_csv, summary_xlsx


def print_basic_findings(threat_df, summary_df):
    print()
    print("Threat model analysis complete.")
    print(f"Total evidence rows analyzed: {len(threat_df)}")

    if threat_df.empty:
        print("No threat rows were generated.")
        return

    print()
    print("Risk level counts:")
    print(threat_df["Risk_Level"].value_counts().to_string())

    print()
    print("Manual review counts:")
    print(threat_df["Needs_Manual_Review"].value_counts().to_string())

    print()
    print("Top 5 category findings:")
    print(threat_df["Category"].value_counts().head(5).to_string())

    if not summary_df.empty:
        print()
        print("Top 5 summarized threat groups:")
        print(summary_df.head(5).to_string(index=False))


def main():
    args = parse_args()

    input_path = Path(args.input)

    print(f"Reading Project 1 output from: {input_path}")

    df = read_input_file(input_path)

    threat_df = analyze_threats(df)
    summary_df = build_summary(threat_df)

    threat_csv, threat_xlsx, summary_csv, summary_xlsx = save_outputs(threat_df, summary_df)

    print_basic_findings(threat_df, summary_df)

    print()
    print(f"Threat model CSV saved to: {threat_csv}")
    print(f"Threat model Excel saved to: {threat_xlsx}")
    print(f"Threat summary CSV saved to: {summary_csv}")
    print(f"Threat summary Excel saved to: {summary_xlsx}")


if __name__ == "__main__":
    main()