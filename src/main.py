"""
LLM Privacy Data Extractor.

Default behavior:
    Shows an interactive CLI menu.
    User can select URLs from config/sources.csv or add custom URLs.
    Scrapes selected URLs.
    Saves raw text into data/raw/.
    Extracts privacy-related keyword matches.
    Saves results into data/output/.
"""

from pathlib import Path
import argparse
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from extractor import extract_privacy_data
from file_utils import read_sources_csv, save_summary
from scraper import scrape_source_to_raw_file


PROJECT_ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = PROJECT_ROOT / "config"
RAW_DATA_DIR = PROJECT_ROOT / "data" / "raw"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output"

SOURCES_CSV = CONFIG_DIR / "sources.csv"

MASTER_DATA_COLUMNS = [
    "Platform",
    "Source_Name",
    "Source URL",
    "Source_Type",
    "Evidence_Quote",
    "Data_Type",
    "Category",
    "Collection_Type",
    "Consumer_or_Api",
    "Purpose_Stated",
    "Retention_Stated",
    "Training_Use_Stated",
    "Third_Party_Sharing_Stated",
    "Disclosure_Quality",
    "Sensitivity",
    "Notes",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract privacy-related data from LLM platform privacy policies and documentation."
    )

    parser.add_argument(
        "--url",
        help="Single privacy policy or documentation URL to scrape."
    )

    parser.add_argument(
        "--platform",
        default="Unknown Platform",
        help="Platform name for single URL mode. Example: Claude"
    )

    parser.add_argument(
        "--source-name",
        default="Unknown Source",
        help='Source name for single URL mode. Example: "Anthropic Privacy Policy"'
    )

    parser.add_argument(
        "--all",
        action="store_true",
        help="Run all URLs from config/sources.csv without showing the interactive menu."
    )

    return parser.parse_args()


def build_sources_from_args(args):
    if args.url:
        return [
            {
                "platform": args.platform,
                "source_name": args.source_name,
                "url": args.url,
            }
        ]

    return None


def print_source_menu(sources):
    print()
    print("Available saved sources:")
    print("-" * 80)

    for index, source in enumerate(sources, start=1):
        print(f"[{index}] {source['platform']} - {source['source_name']}")
        print(f"    {source['url']}")

    print("-" * 80)
    print()


def parse_selection(selection_text, max_number):
    selection_text = selection_text.strip().lower()

    if selection_text == "all":
        return list(range(max_number))

    if selection_text in {"none", "n", "no", ""}:
        return []

    selected_indexes = set()

    for part in selection_text.split(","):
        part = part.strip()

        if not part:
            continue

        if "-" in part:
            start_text, end_text = part.split("-", 1)

            if not start_text.strip().isdigit() or not end_text.strip().isdigit():
                print(f"Invalid range ignored: {part}")
                continue

            start = int(start_text.strip())
            end = int(end_text.strip())

            if start > end:
                print(f"Invalid range ignored: {part}")
                continue

            for number in range(start, end + 1):
                if 1 <= number <= max_number:
                    selected_indexes.add(number - 1)
                else:
                    print(f"Number out of range ignored: {number}")

        else:
            if not part.isdigit():
                print(f"Invalid selection ignored: {part}")
                continue

            number = int(part)

            if 1 <= number <= max_number:
                selected_indexes.add(number - 1)
            else:
                print(f"Number out of range ignored: {number}")

    return sorted(selected_indexes)


def ask_yes_no(prompt):
    while True:
        answer = input(prompt).strip().lower()

        if answer in {"y", "yes"}:
            return True

        if answer in {"n", "no"}:
            return False

        print("Please enter y or n.")


def collect_custom_sources():
    custom_sources = []

    print()
    print("Custom URL entry")
    print()

    while True:
        url = input("Enter custom URL: ").strip()

        if not url:
            print("No URL entered. Returning to source selection.")
            break

        platform = input("Platform name, example Claude/OpenAI/Gemini/Copilot: ").strip()
        source_name = input("Source name, example Privacy Policy/API Docs: ").strip()

        if not platform:
            platform = "Custom Platform"

        if not source_name:
            source_name = "Custom Source"

        custom_sources.append({
            "platform": platform,
            "source_name": source_name,
            "url": url,
        })

        print("Custom source added.")
        print()

        if not ask_yes_no("Add another custom URL? (y/n): "):
            break

        print()

    return custom_sources


def interactive_select_sources(saved_sources):
    print_source_menu(saved_sources)

    print("Select sources to scrape.")
    print("Examples:")
    print("  1")
    print("  1,3,5")
    print("  1-4")
    print("  1,3-5")
    print("  all")
    print("  none")
    print()

    selection_text = input("Enter your selection: ")

    selected_indexes = parse_selection(
        selection_text=selection_text,
        max_number=len(saved_sources),
    )

    selected_sources = [saved_sources[index] for index in selected_indexes]

    print()
    print(f"Saved sources selected: {len(selected_sources)}")

    if ask_yes_no("Add custom URL(s)? (y/n): "):
        selected_sources.extend(collect_custom_sources())

    print()
    print("Final selected sources:")
    print("-" * 80)

    if not selected_sources:
        print("No sources selected.")
    else:
        for index, source in enumerate(selected_sources, start=1):
            print(f"[{index}] {source['platform']} - {source['source_name']}")
            print(f"    {source['url']}")

    print("-" * 80)
    print()

    if not selected_sources:
        return []

    if not ask_yes_no("Continue with these sources? (y/n): "):
        print("Canceled by user.")
        return []

    return selected_sources


def infer_source_type(source_name, source_url):
    text = f"{source_name} {source_url}".lower()

    if "privacy" in text or "policy" in text:
        return "Policy"

    if "api" in text or "docs" in text or "documentation" in text:
        return "API Docs"

    if "terms" in text:
        return "Terms"

    if "cookie" in text:
        return "Cookie Notice"

    if "support" in text or "help" in text:
        return "Support"

    return "Public documentation"


def infer_collection_type(category):
    if category in {"Account Information", "Payment Information", "User Content"}:
        return "Direct"

    if category in {"Device and Technical Data", "Location Data", "Cookies and Tracking"}:
        return "Automatic"

    if category == "Third Party Sharing":
        return "Sharing / disclosure"

    if category == "Retention and Storage":
        return "Retention / storage"

    if category == "Model Improvement and Training":
        return "Training / improvement"

    if category == "Security and Safety Data":
        return "Security / safety"

    return "Needs manual review"


def infer_consumer_or_api(source_name, source_url):
    text = f"{source_name} {source_url}".lower()

    if "api" in text or "docs" in text or "commercial" in text:
        return "API / commercial contexts"

    if "consumer" in text:
        return "Consumer contexts"

    return "Consumer / platform contexts"


def infer_sensitivity(category):
    if category in {
        "Account Information",
        "Payment Information",
        "User Content",
        "Location Data",
        "Security and Safety Data",
    }:
        return "Medium/High"

    if category in {
        "Device and Technical Data",
        "Cookies and Tracking",
        "Third Party Sharing",
        "Model Improvement and Training",
        "Retention and Storage",
    }:
        return "Medium"

    return "Needs manual review"


def build_purpose_stated(category, sentence):
    if category in {
        "Account Information",
        "Payment Information",
        "User Content",
        "Device and Technical Data",
        "Location Data",
        "Cookies and Tracking",
        "Security and Safety Data",
    }:
        return sentence

    return None


def build_retention_stated(category, sentence):
    if category == "Retention and Storage":
        return sentence

    return None


def build_training_use_stated(category, sentence):
    if category == "Model Improvement and Training":
        return sentence

    return "No / not stated in row"


def build_third_party_sharing_stated(category, sentence):
    if category == "Third Party Sharing":
        return sentence

    return "No / not stated in row"


def format_results_for_master_data(raw_results):
    formatted_rows = []

    for row in raw_results:
        category = row.get("data_category")
        sentence = row.get("matching_sentence")

        formatted_rows.append({
            "Platform": row.get("platform"),
            "Source_Name": row.get("source_name"),
            "Source URL": row.get("source_url"),
            "Source_Type": infer_source_type(
                row.get("source_name", ""),
                row.get("source_url", ""),
            ),
            "Evidence_Quote": sentence,
            "Data_Type": row.get("keyword"),
            "Category": category,
            "Collection_Type": infer_collection_type(category),
            "Consumer_or_Api": infer_consumer_or_api(
                row.get("source_name", ""),
                row.get("source_url", ""),
            ),
            "Purpose_Stated": build_purpose_stated(category, sentence),
            "Retention_Stated": build_retention_stated(category, sentence),
            "Training_Use_Stated": build_training_use_stated(category, sentence),
            "Third_Party_Sharing_Stated": build_third_party_sharing_stated(category, sentence),
            "Disclosure_Quality": "Extracted keyword match - needs manual review",
            "Sensitivity": infer_sensitivity(category),
            "Notes": f"Matched keyword: {row.get('keyword')}",
        })

    return formatted_rows


def format_excel_file(excel_path):
    """
    Applies readable formatting to the Excel output:
        - wider columns
        - wrapped headers
        - wrapped evidence text
        - frozen header row
        - auto filter
    """
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    worksheet.title = "Master_Data"

    column_widths = {
        "A": 18,  # Platform
        "B": 34,  # Source_Name
        "C": 55,  # Source URL
        "D": 22,  # Source_Type
        "E": 90,  # Evidence_Quote
        "F": 26,  # Data_Type
        "G": 30,  # Category
        "H": 28,  # Collection_Type
        "I": 30,  # Consumer_or_Api
        "J": 60,  # Purpose_Stated
        "K": 60,  # Retention_Stated
        "L": 60,  # Training_Use_Stated
        "M": 60,  # Third_Party_Sharing_Stated
        "N": 42,  # Disclosure_Quality
        "O": 20,  # Sensitivity
        "P": 38,  # Notes
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # fix header
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 45

    # format rows so its not bunched up
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(excel_path)


def save_master_data_output(formatted_results):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    csv_path = OUTPUT_DIR / "master_data_privacy_output.csv"
    excel_path = OUTPUT_DIR / "master_data_privacy_output.xlsx"

    df = pd.DataFrame(formatted_results, columns=MASTER_DATA_COLUMNS)

    df.to_csv(csv_path, index=False)
    df.to_excel(excel_path, index=False)

    format_excel_file(excel_path)

    return df, csv_path, excel_path


def run_extraction(sources):
    all_results = []

    for source in sources:
        try:
            raw_file_path, text = scrape_source_to_raw_file(
                source=source,
                raw_data_dir=RAW_DATA_DIR,
            )
        except Exception as error:
            print()
            print(f"Failed to scrape source: {source.get('source_name')}")
            print(f"URL: {source.get('url')}")
            print(f"Error: {error}")
            print("Skipping this source.")
            print()
            continue

        results = extract_privacy_data(
            platform=source["platform"],
            source_name=source["source_name"],
            source_url=source["url"],
            source_file=str(raw_file_path.relative_to(PROJECT_ROOT)),
            text=text,
        )

        print(f"Matches found: {len(results)}")
        print()

        all_results.extend(results)

    formatted_results = format_results_for_master_data(all_results)
    master_df, master_csv_path, master_excel_path = save_master_data_output(formatted_results)

    raw_df = pd.DataFrame(all_results)

    summary_csv_path = OUTPUT_DIR / "extraction_summary.csv"
    summary_excel_path = OUTPUT_DIR / "extraction_summary.xlsx"

    summary_df = save_summary(
        df=raw_df,
        csv_path=summary_csv_path,
        excel_path=summary_excel_path,
    )

    print("Extraction complete.")
    print(f"Total rows extracted: {len(master_df)}")
    print(f"Master-format CSV saved to: {master_csv_path}")
    print(f"Master-format Excel saved to: {master_excel_path}")
    print(f"Summary CSV saved to: {summary_csv_path}")
    print(f"Summary Excel saved to: {summary_excel_path}")
    print(f"Summary rows: {len(summary_df)}")


def main():
    args = parse_args()

    cli_sources = build_sources_from_args(args)

    if cli_sources:
        print("Running in single URL mode.")
        sources = cli_sources

    else:
        print(f"Reading saved source database from: {SOURCES_CSV}")
        saved_sources = read_sources_csv(SOURCES_CSV)

        if args.all:
            print("Running all saved sources.")
            sources = saved_sources
        else:
            sources = interactive_select_sources(saved_sources)

    if not sources:
        print("No sources to process. Exiting.")
        return

    run_extraction(sources)


if __name__ == "__main__":
    main()